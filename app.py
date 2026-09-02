import os
import math
import sqlite3
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, flash, session, g, send_file
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "changez-moi-en-production")

APP_PASSWORD = os.environ.get("APP_PASSWORD", "buvette")
ASSOCIATION_NOM = os.environ.get("ASSOCIATION_NOM", "Mon association")

DB_PATH = os.path.join(os.path.dirname(__file__), "buvette.db")
UPLOAD_DIR = os.path.join(app.static_folder, "uploads")
ALLOWED_IMAGE_EXT = {"png", "jpg", "jpeg", "webp"}

app.jinja_env.globals["association_nom"] = ASSOCIATION_NOM


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = get_db()
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS boissons (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            contenant TEXT,
            servi TEXT,
            prix_vente REAL DEFAULT 0,
            qt_par_pack INTEGER DEFAULT 1,
            fournisseur TEXT,
            commentaire TEXT,
            ordre INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS snacks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            servi TEXT,
            prix_vente REAL DEFAULT 0,
            commentaire TEXT,
            ordre INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS snack_composants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            snack_id INTEGER NOT NULL,
            nom TEXT NOT NULL,
            quantite REAL DEFAULT 0,
            unite TEXT,
            qt_par_pack INTEGER DEFAULT 1,
            fournisseur TEXT,
            FOREIGN KEY (snack_id) REFERENCES snacks(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS competitions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            date TEXT,
            titre_affiche TEXT,
            titre_affiche_snacks TEXT
        );
        CREATE TABLE IF NOT EXISTS selections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            competition_id INTEGER NOT NULL,
            boisson_id INTEGER NOT NULL,
            quantite INTEGER NOT NULL,
            UNIQUE(competition_id, boisson_id),
            FOREIGN KEY (competition_id) REFERENCES competitions(id) ON DELETE CASCADE,
            FOREIGN KEY (boisson_id) REFERENCES boissons(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS snack_selections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            competition_id INTEGER NOT NULL,
            snack_id INTEGER NOT NULL,
            quantite INTEGER NOT NULL,
            UNIQUE(competition_id, snack_id),
            FOREIGN KEY (competition_id) REFERENCES competitions(id) ON DELETE CASCADE,
            FOREIGN KEY (snack_id) REFERENCES snacks(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS affiches (
            competition_id INTEGER PRIMARY KEY,
            background_filename TEXT,
            FOREIGN KEY (competition_id) REFERENCES competitions(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS affiches_snacks (
            competition_id INTEGER PRIMARY KEY,
            background_filename TEXT,
            FOREIGN KEY (competition_id) REFERENCES competitions(id) ON DELETE CASCADE
        );
        """
    )
    db.commit()


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("authenticated"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if request.form.get("password") == APP_PASSWORD:
            session["authenticated"] = True
            return redirect(request.args.get("next") or url_for("boissons"))
        flash("Mot de passe incorrect.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


def distinct_values(table, column):
    rows = get_db().execute(
        f"SELECT DISTINCT {column} FROM {table} WHERE {column} IS NOT NULL AND {column} != '' ORDER BY {column}"
    ).fetchall()
    return [r[0] for r in rows]


def get_competition_or_404(comp_id):
    comp = get_db().execute("SELECT * FROM competitions WHERE id = ?", (comp_id,)).fetchone()
    if comp is None:
        from flask import abort
        abort(404)
    return comp


@app.route("/")
@login_required
def index():
    return redirect(url_for("boissons"))


# ---------------------------------------------------------------- Boissons

@app.route("/boissons")
@login_required
def boissons():
    db = get_db()
    comp_id = request.args.get("competition_id", type=int)
    if comp_id:
        rows = db.execute(
            """
            SELECT b.* FROM boissons b
            JOIN selections s ON s.boisson_id = b.id
            WHERE s.competition_id = ? AND s.quantite > 0
            ORDER BY b.ordre, b.nom
            """,
            (comp_id,),
        ).fetchall()
    else:
        rows = db.execute("SELECT * FROM boissons ORDER BY ordre, nom").fetchall()
    competitions_list = db.execute("SELECT * FROM competitions ORDER BY date DESC, nom").fetchall()
    return render_template(
        "boissons.html",
        boissons=rows,
        competitions=competitions_list,
        selected_competition_id=comp_id,
        contenants=distinct_values("boissons", "contenant"),
        servis=distinct_values("boissons", "servi"),
        fournisseurs=distinct_values("boissons", "fournisseur"),
        boisson_noms=distinct_values("boissons", "nom"),
    )


@app.route("/boissons/ajouter", methods=["POST"])
@login_required
def boissons_ajouter():
    nom = request.form.get("nom", "").strip()
    comp_id = request.form.get("competition_id", type=int)
    if nom:
        db = get_db()
        max_ordre = db.execute("SELECT COALESCE(MAX(ordre), 0) FROM boissons").fetchone()[0]
        cur = db.execute(
            "INSERT INTO boissons (nom, contenant, servi, prix_vente, qt_par_pack, fournisseur, commentaire, ordre) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                nom,
                request.form.get("contenant", "").strip(),
                request.form.get("servi", "").strip(),
                float(request.form.get("prix_vente") or 0),
                int(request.form.get("qt_par_pack") or 1),
                request.form.get("fournisseur", "").strip(),
                request.form.get("commentaire", "").strip(),
                max_ordre + 1,
            ),
        )
        if comp_id:
            db.execute(
                "INSERT INTO selections (competition_id, boisson_id, quantite) VALUES (?, ?, 1)",
                (comp_id, cur.lastrowid),
            )
        db.commit()
    return redirect(url_for("boissons", competition_id=comp_id) if comp_id else url_for("boissons"))


@app.route("/boissons/<int:boisson_id>/modifier", methods=["POST"])
@login_required
def boissons_modifier(boisson_id):
    db = get_db()
    db.execute(
        "UPDATE boissons SET nom=?, contenant=?, servi=?, prix_vente=?, qt_par_pack=?, fournisseur=?, commentaire=? WHERE id=?",
        (
            request.form.get("nom", "").strip(),
            request.form.get("contenant", "").strip(),
            request.form.get("servi", "").strip(),
            float(request.form.get("prix_vente") or 0),
            int(request.form.get("qt_par_pack") or 1),
            request.form.get("fournisseur", "").strip(),
            request.form.get("commentaire", "").strip(),
            boisson_id,
        ),
    )
    db.commit()
    comp_id = request.form.get("competition_id", type=int)
    return redirect(url_for("boissons", competition_id=comp_id) if comp_id else url_for("boissons"))


@app.route("/boissons/<int:boisson_id>/supprimer", methods=["POST"])
@login_required
def boissons_supprimer(boisson_id):
    db = get_db()
    db.execute("DELETE FROM boissons WHERE id = ?", (boisson_id,))
    db.commit()
    return redirect(url_for("boissons"))


def move_within_view(db, table, item_id, direction, visible_ids):
    """Déplace item_id d'une position par rapport à ses voisins VISIBLES
    (visible_ids, dans l'ordre d'affichage courant), tout en réinsérant
    l'élément au bon endroit dans l'ordre global (les éléments non visibles
    gardent leur position relative)."""
    if item_id not in visible_ids:
        return
    pos = visible_ids.index(item_id)
    target_pos = pos - 1 if direction == "up" else pos + 1
    if target_pos < 0 or target_pos >= len(visible_ids):
        return
    neighbor_id = visible_ids[target_pos]

    full_rows = db.execute(f"SELECT id FROM {table} ORDER BY ordre, nom").fetchall()
    full_ids = [r["id"] for r in full_rows]
    if item_id not in full_ids or neighbor_id not in full_ids:
        return
    full_ids.remove(item_id)
    idx_neighbor = full_ids.index(neighbor_id)
    insert_at = idx_neighbor if direction == "up" else idx_neighbor + 1
    full_ids.insert(insert_at, item_id)

    for i, fid in enumerate(full_ids):
        db.execute(f"UPDATE {table} SET ordre = ? WHERE id = ?", (i, fid))
    db.commit()


@app.route("/boissons/<int:boisson_id>/deplacer/<direction>", methods=["POST"])
@login_required
def boissons_deplacer(boisson_id, direction):
    db = get_db()
    comp_id = request.form.get("competition_id", type=int)
    if comp_id:
        visible_rows = db.execute(
            """
            SELECT b.id FROM boissons b
            JOIN selections s ON s.boisson_id = b.id
            WHERE s.competition_id = ? AND s.quantite > 0
            ORDER BY b.ordre, b.nom
            """,
            (comp_id,),
        ).fetchall()
    else:
        visible_rows = db.execute("SELECT id FROM boissons ORDER BY ordre, nom").fetchall()
    visible_ids = [r["id"] for r in visible_rows]
    move_within_view(db, "boissons", boisson_id, direction, visible_ids)
    return redirect(url_for("boissons", competition_id=comp_id) if comp_id else url_for("boissons"))


# ------------------------------------------------------------------- Snacks

@app.route("/snacks")
@login_required
def snacks():
    db = get_db()
    comp_id = request.args.get("competition_id", type=int)
    if comp_id:
        rows = db.execute(
            """
            SELECT sn.* FROM snacks sn
            JOIN snack_selections ss ON ss.snack_id = sn.id
            WHERE ss.competition_id = ? AND ss.quantite > 0
            ORDER BY sn.ordre, sn.nom
            """,
            (comp_id,),
        ).fetchall()
    else:
        rows = db.execute("SELECT * FROM snacks ORDER BY ordre, nom").fetchall()
    snack_ids = [r["id"] for r in rows]
    composants_by_snack = {}
    if snack_ids:
        placeholders = ",".join("?" * len(snack_ids))
        comp_rows = db.execute(
            f"SELECT * FROM snack_composants WHERE snack_id IN ({placeholders}) ORDER BY id", snack_ids
        ).fetchall()
        for c in comp_rows:
            composants_by_snack.setdefault(c["snack_id"], []).append(c)
    competitions_list = db.execute("SELECT * FROM competitions ORDER BY date DESC, nom").fetchall()
    return render_template(
        "snacks.html",
        snacks=rows,
        composants_by_snack=composants_by_snack,
        competitions=competitions_list,
        selected_competition_id=comp_id,
        snack_noms=distinct_values("snacks", "nom"),
        snack_servis=distinct_values("snacks", "servi"),
        composant_noms=distinct_values("snack_composants", "nom"),
        composant_unites=distinct_values("snack_composants", "unite"),
        composant_fournisseurs=distinct_values("snack_composants", "fournisseur"),
    )


@app.route("/snacks/ajouter", methods=["POST"])
@login_required
def snacks_ajouter():
    nom = request.form.get("nom", "").strip()
    comp_id = request.form.get("competition_id", type=int)
    if nom:
        db = get_db()
        max_ordre = db.execute("SELECT COALESCE(MAX(ordre), 0) FROM snacks").fetchone()[0]
        cur = db.execute(
            "INSERT INTO snacks (nom, servi, prix_vente, commentaire, ordre) VALUES (?, ?, ?, ?, ?)",
            (
                nom,
                request.form.get("servi", "").strip(),
                float(request.form.get("prix_vente") or 0),
                request.form.get("commentaire", "").strip(),
                max_ordre + 1,
            ),
        )
        if comp_id:
            db.execute(
                "INSERT INTO snack_selections (competition_id, snack_id, quantite) VALUES (?, ?, 1)",
                (comp_id, cur.lastrowid),
            )
        db.commit()
    return redirect(url_for("snacks", competition_id=comp_id) if comp_id else url_for("snacks"))


@app.route("/snacks/<int:snack_id>/modifier", methods=["POST"])
@login_required
def snacks_modifier(snack_id):
    db = get_db()
    db.execute(
        "UPDATE snacks SET nom=?, servi=?, prix_vente=?, commentaire=? WHERE id=?",
        (
            request.form.get("nom", "").strip(),
            request.form.get("servi", "").strip(),
            float(request.form.get("prix_vente") or 0),
            request.form.get("commentaire", "").strip(),
            snack_id,
        ),
    )
    db.commit()
    comp_id = request.form.get("competition_id", type=int)
    return redirect(url_for("snacks", competition_id=comp_id) if comp_id else url_for("snacks"))


@app.route("/snacks/<int:snack_id>/supprimer", methods=["POST"])
@login_required
def snacks_supprimer(snack_id):
    db = get_db()
    db.execute("DELETE FROM snacks WHERE id = ?", (snack_id,))
    db.commit()
    return redirect(url_for("snacks"))


@app.route("/snacks/<int:snack_id>/deplacer/<direction>", methods=["POST"])
@login_required
def snacks_deplacer(snack_id, direction):
    db = get_db()
    comp_id = request.form.get("competition_id", type=int)
    if comp_id:
        visible_rows = db.execute(
            """
            SELECT sn.id FROM snacks sn
            JOIN snack_selections ss ON ss.snack_id = sn.id
            WHERE ss.competition_id = ? AND ss.quantite > 0
            ORDER BY sn.ordre, sn.nom
            """,
            (comp_id,),
        ).fetchall()
    else:
        visible_rows = db.execute("SELECT id FROM snacks ORDER BY ordre, nom").fetchall()
    visible_ids = [r["id"] for r in visible_rows]
    move_within_view(db, "snacks", snack_id, direction, visible_ids)
    return redirect(url_for("snacks", competition_id=comp_id) if comp_id else url_for("snacks"))


@app.route("/snacks/<int:snack_id>/composants/ajouter", methods=["POST"])
@login_required
def composants_ajouter(snack_id):
    nom = request.form.get("nom", "").strip()
    if nom:
        db = get_db()
        db.execute(
            "INSERT INTO snack_composants (snack_id, nom, quantite, unite, qt_par_pack, fournisseur) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                snack_id,
                nom,
                float(request.form.get("quantite") or 0),
                request.form.get("unite", "").strip(),
                int(request.form.get("qt_par_pack") or 1),
                request.form.get("fournisseur", "").strip(),
            ),
        )
        db.commit()
    return redirect(url_for("snacks"))


@app.route("/composants/<int:composant_id>/modifier", methods=["POST"])
@login_required
def composants_modifier(composant_id):
    db = get_db()
    db.execute(
        "UPDATE snack_composants SET nom=?, quantite=?, unite=?, qt_par_pack=?, fournisseur=? WHERE id=?",
        (
            request.form.get("nom", "").strip(),
            float(request.form.get("quantite") or 0),
            request.form.get("unite", "").strip(),
            int(request.form.get("qt_par_pack") or 1),
            request.form.get("fournisseur", "").strip(),
            composant_id,
        ),
    )
    db.commit()
    return redirect(url_for("snacks"))


@app.route("/composants/<int:composant_id>/supprimer", methods=["POST"])
@login_required
def composants_supprimer(composant_id):
    db = get_db()
    db.execute("DELETE FROM snack_composants WHERE id = ?", (composant_id,))
    db.commit()
    return redirect(url_for("snacks"))


# ------------------------------------------------------------ Competitions

@app.route("/competitions")
@login_required
def competitions():
    db = get_db()
    comps = db.execute("SELECT * FROM competitions ORDER BY date DESC, nom").fetchall()
    return render_template("competitions.html", competitions=comps)


@app.route("/competitions/creer", methods=["POST"])
@login_required
def competitions_creer():
    nom = request.form.get("nom", "").strip()
    if nom:
        db = get_db()
        cur = db.execute(
            "INSERT INTO competitions (nom, date, titre_affiche, titre_affiche_snacks) VALUES (?, ?, '', '')",
            (nom, request.form.get("date", "").strip()),
        )
        db.commit()
        return redirect(url_for("competition_detail", comp_id=cur.lastrowid))
    return redirect(url_for("competitions"))


@app.route("/competitions/<int:comp_id>/supprimer", methods=["POST"])
@login_required
def competitions_supprimer(comp_id):
    db = get_db()
    db.execute("DELETE FROM competitions WHERE id = ?", (comp_id,))
    db.commit()
    return redirect(url_for("competitions"))


@app.route("/competitions/<int:comp_id>")
@login_required
def competition_detail(comp_id):
    db = get_db()
    comp = get_competition_or_404(comp_id)
    drinks = db.execute("SELECT * FROM boissons ORDER BY ordre, nom").fetchall()
    drink_selections = {
        r["boisson_id"]: r["quantite"]
        for r in db.execute(
            "SELECT boisson_id, quantite FROM selections WHERE competition_id = ?", (comp_id,)
        ).fetchall()
    }
    snacks_list = db.execute("SELECT * FROM snacks ORDER BY ordre, nom").fetchall()
    snack_selections = {
        r["snack_id"]: r["quantite"]
        for r in db.execute(
            "SELECT snack_id, quantite FROM snack_selections WHERE competition_id = ?", (comp_id,)
        ).fetchall()
    }
    return render_template(
        "competition_detail.html",
        competition=comp,
        boissons=drinks,
        drink_selections=drink_selections,
        snacks=snacks_list,
        snack_selections=snack_selections,
    )


@app.route("/competitions/<int:comp_id>/titre", methods=["POST"])
@login_required
def competition_titre(comp_id):
    get_competition_or_404(comp_id)
    db = get_db()
    db.execute(
        "UPDATE competitions SET titre_affiche = ?, titre_affiche_snacks = ? WHERE id = ?",
        (
            request.form.get("titre_affiche", "").strip(),
            request.form.get("titre_affiche_snacks", "").strip(),
            comp_id,
        ),
    )
    db.commit()
    return redirect(url_for("competition_detail", comp_id=comp_id))


@app.route("/competitions/<int:comp_id>/quantites", methods=["POST"])
@login_required
def competition_quantites(comp_id):
    db = get_db()
    get_competition_or_404(comp_id)
    for b in db.execute("SELECT id FROM boissons").fetchall():
        raw = request.form.get(f"boisson_{b['id']}", "").strip()
        qty = int(raw) if raw.isdigit() else 0
        if qty > 0:
            db.execute(
                "INSERT INTO selections (competition_id, boisson_id, quantite) VALUES (?, ?, ?) "
                "ON CONFLICT(competition_id, boisson_id) DO UPDATE SET quantite = excluded.quantite",
                (comp_id, b["id"], qty),
            )
        else:
            db.execute("DELETE FROM selections WHERE competition_id = ? AND boisson_id = ?", (comp_id, b["id"]))
    for s in db.execute("SELECT id FROM snacks").fetchall():
        raw = request.form.get(f"snack_{s['id']}", "").strip()
        qty = int(raw) if raw.isdigit() else 0
        if qty > 0:
            db.execute(
                "INSERT INTO snack_selections (competition_id, snack_id, quantite) VALUES (?, ?, ?) "
                "ON CONFLICT(competition_id, snack_id) DO UPDATE SET quantite = excluded.quantite",
                (comp_id, s["id"], qty),
            )
        else:
            db.execute("DELETE FROM snack_selections WHERE competition_id = ? AND snack_id = ?", (comp_id, s["id"]))
    db.commit()
    flash("Quantités mises à jour.", "success")
    return redirect(url_for("competition_detail", comp_id=comp_id))


# ------------------------------------------------------------------ Commande

@app.route("/competitions/<int:comp_id>/commande")
@login_required
def commande(comp_id):
    db = get_db()
    comp = get_competition_or_404(comp_id)
    rows = db.execute(
        """
        SELECT b.nom, b.contenant, b.qt_par_pack, b.fournisseur, s.quantite
        FROM selections s JOIN boissons b ON b.id = s.boisson_id
        WHERE s.competition_id = ? AND s.quantite > 0 ORDER BY b.fournisseur, b.nom
        """,
        (comp_id,),
    ).fetchall()
    par_fournisseur = {}
    for r in rows:
        fournisseur = r["fournisseur"] or "Fournisseur non renseigné"
        pack = r["qt_par_pack"] or 1
        packs = math.ceil(r["quantite"] / pack) if pack else r["quantite"]
        par_fournisseur.setdefault(fournisseur, []).append(
            {"nom": r["nom"], "contenant": r["contenant"], "quantite": r["quantite"], "qt_par_pack": pack, "packs": packs}
        )
    return render_template("commande.html", competition=comp, par_fournisseur=par_fournisseur)


@app.route("/competitions/<int:comp_id>/commande/export")
@login_required
def commande_export(comp_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO

    db = get_db()
    comp = get_competition_or_404(comp_id)
    rows = db.execute(
        """
        SELECT b.nom, b.contenant, b.qt_par_pack, b.fournisseur, s.quantite
        FROM selections s JOIN boissons b ON b.id = s.boisson_id
        WHERE s.competition_id = ? AND s.quantite > 0 ORDER BY b.fournisseur, b.nom
        """,
        (comp_id,),
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commande"
    headers = ["Fournisseur", "Boisson", "Contenant", "Quantité totale", "Qté / pack", "Packs à commander"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="14524A")
    for r in rows:
        pack = r["qt_par_pack"] or 1
        packs = math.ceil(r["quantite"] / pack) if pack else r["quantite"]
        ws.append([r["fournisseur"] or "Fournisseur non renseigné", r["nom"], r["contenant"] or "", r["quantite"], pack, packs])
    for i, w in enumerate([22, 22, 18, 16, 12, 18], start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Commande - {comp['nom']}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/competitions/<int:comp_id>/commande-snacks")
@login_required
def commande_snacks(comp_id):
    db = get_db()
    comp = get_competition_or_404(comp_id)
    par_fournisseur = _compute_snack_order(db, comp_id)
    return render_template("commande_snacks.html", competition=comp, par_fournisseur=par_fournisseur)


def _compute_snack_order(db, comp_id):
    rows = db.execute(
        """
        SELECT sn.id as snack_id, ss.quantite as snack_qty
        FROM snack_selections ss JOIN snacks sn ON sn.id = ss.snack_id
        WHERE ss.competition_id = ? AND ss.quantite > 0
        """,
        (comp_id,),
    ).fetchall()
    totals = {}
    for r in rows:
        composants = db.execute("SELECT * FROM snack_composants WHERE snack_id = ?", (r["snack_id"],)).fetchall()
        for c in composants:
            key = (c["nom"], c["fournisseur"] or "", c["unite"] or "")
            if key not in totals:
                totals[key] = {
                    "nom": c["nom"],
                    "unite": c["unite"] or "",
                    "fournisseur": c["fournisseur"] or "Fournisseur non renseigné",
                    "qt_par_pack": c["qt_par_pack"] or 1,
                    "total": 0,
                }
            totals[key]["total"] += r["snack_qty"] * (c["quantite"] or 0)
    par_fournisseur = {}
    for t in totals.values():
        packs = math.ceil(t["total"] / t["qt_par_pack"]) if t["qt_par_pack"] else t["total"]
        par_fournisseur.setdefault(t["fournisseur"], []).append(
            {"nom": t["nom"], "unite": t["unite"], "total": t["total"], "qt_par_pack": t["qt_par_pack"], "packs": packs}
        )
    return par_fournisseur


@app.route("/competitions/<int:comp_id>/commande-snacks/export")
@login_required
def commande_snacks_export(comp_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO

    db = get_db()
    comp = get_competition_or_404(comp_id)
    par_fournisseur = _compute_snack_order(db, comp_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commande snacks"
    headers = ["Fournisseur", "Composant", "Quantité totale", "Unité", "Qté / pack", "Packs à commander"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="14524A")
    for fournisseur, items in par_fournisseur.items():
        for it in items:
            ws.append([fournisseur, it["nom"], it["total"], it["unite"], it["qt_par_pack"], it["packs"]])
    for i, w in enumerate([22, 22, 16, 12, 12, 18], start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Commande snacks - {comp['nom']}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# -------------------------------------------------------------------- Affiche

@app.route("/competitions/<int:comp_id>/affiche")
@login_required
def affiche(comp_id):
    db = get_db()
    comp = get_competition_or_404(comp_id)
    rows = db.execute(
        """
        SELECT b.nom, b.contenant, b.servi, b.prix_vente
        FROM selections s JOIN boissons b ON b.id = s.boisson_id
        WHERE s.competition_id = ? AND s.quantite > 0 ORDER BY b.ordre, b.nom
        """,
        (comp_id,),
    ).fetchall()
    fond = db.execute("SELECT background_filename FROM affiches WHERE competition_id = ?", (comp_id,)).fetchone()
    return render_template(
        "affiche.html", competition=comp, boissons=rows,
        fond_filename=fond["background_filename"] if fond else None,
    )


@app.route("/competitions/<int:comp_id>/affiche/fond", methods=["POST"])
@login_required
def affiche_fond(comp_id):
    get_competition_or_404(comp_id)
    fichier = request.files.get("fond")
    if fichier and fichier.filename:
        ext = fichier.filename.rsplit(".", 1)[-1].lower()
        if ext in ALLOWED_IMAGE_EXT:
            os.makedirs(UPLOAD_DIR, exist_ok=True)
            filename = secure_filename(f"fond_{comp_id}.{ext}")
            fichier.save(os.path.join(UPLOAD_DIR, filename))
            db = get_db()
            db.execute(
                "INSERT INTO affiches (competition_id, background_filename) VALUES (?, ?) "
                "ON CONFLICT(competition_id) DO UPDATE SET background_filename = excluded.background_filename",
                (comp_id, filename),
            )
            db.commit()
        else:
            flash("Format d'image non supporté (png, jpg, jpeg, webp).", "error")
    return redirect(url_for("affiche", comp_id=comp_id))


@app.route("/competitions/<int:comp_id>/affiche-snacks")
@login_required
def affiche_snacks(comp_id):
    db = get_db()
    comp = get_competition_or_404(comp_id)
    rows = db.execute(
        """
        SELECT sn.nom, sn.servi, sn.prix_vente
        FROM snack_selections ss JOIN snacks sn ON sn.id = ss.snack_id
        WHERE ss.competition_id = ? AND ss.quantite > 0 ORDER BY sn.ordre, sn.nom
        """,
        (comp_id,),
    ).fetchall()
    fond = db.execute("SELECT background_filename FROM affiches_snacks WHERE competition_id = ?", (comp_id,)).fetchone()
    return render_template(
        "affiche_snacks.html", competition=comp, snacks=rows,
        fond_filename=fond["background_filename"] if fond else None,
    )


@app.route("/competitions/<int:comp_id>/affiche-snacks/fond", methods=["POST"])
@login_required
def affiche_snacks_fond(comp_id):
    get_competition_or_404(comp_id)
    fichier = request.files.get("fond")
    if fichier and fichier.filename:
        ext = fichier.filename.rsplit(".", 1)[-1].lower()
        if ext in ALLOWED_IMAGE_EXT:
            os.makedirs(UPLOAD_DIR, exist_ok=True)
            filename = secure_filename(f"fond_snacks_{comp_id}.{ext}")
            fichier.save(os.path.join(UPLOAD_DIR, filename))
            db = get_db()
            db.execute(
                "INSERT INTO affiches_snacks (competition_id, background_filename) VALUES (?, ?) "
                "ON CONFLICT(competition_id) DO UPDATE SET background_filename = excluded.background_filename",
                (comp_id, filename),
            )
            db.commit()
        else:
            flash("Format d'image non supporté (png, jpg, jpeg, webp).", "error")
    return redirect(url_for("affiche_snacks", comp_id=comp_id))


with app.app_context():
    init_db()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
